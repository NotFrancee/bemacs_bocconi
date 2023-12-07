import openpyxl
import pandas as pd
import numpy as np
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment


def style_table(df: pd.DataFrame, filename: str):
    df.to_excel(filename)

    h_style = NamedStyle("heading")
    h_style.font = Font(bold=True, size=14, color="00FF0000")
    h_style.alignment = Alignment(horizontal="center", vertical="center")
    h_style.fill = PatternFill(patternType=None, bgColor="00000000")  # bsic color here

    wb = openpyxl.load_workbook(filename)
    wb.add_named_style(h_style)

    ws = wb.active

    if ws is None:
        raise Exception("No active worksheet")

    for cell in ws["1:1"]:
        print(cell.value)
        cell.style = "heading"

    wb.save(filename)


# Create example dataframe with random data
data = {
    "A": np.random.randint(0, 100, 5),
    "B": np.random.randint(0, 100, 5),
    "C": np.random.randint(0, 100, 5),
}
df = pd.DataFrame(data)

style_table(df, "test.xlsx")
